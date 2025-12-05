"""
Excel Builder - Professional Excel Report Generation
Agent 18 - Investment-Grade Excel Reports using openpyxl

Creates comprehensive Excel reports with:
- Multiple worksheets for different data views
- Professional formatting and styling
- Charts and visualizations
- Pivot-ready data structures
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List
import logging

logger = logging.getLogger(__name__)


class ExcelBuilder:
    """
    Professional Excel report builder
    Creates multi-worksheet workbooks with charts and formatting
    """

    def __init__(self, output_path: str):
        """Initialize Excel builder with output path"""
        self.output_path = output_path
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

        # Define color scheme
        self.colors = {
            'header_bg': 'FFblue2c5aa0',
            'header_text': 'FFFFFFFF',
            'alt_row': 'FFF5F5F5',
            'positive': 'FF4CAF50',
            'negative': 'FFF44336',
            'warning': 'FFFFC107'
        }

        # Define styles
        self._setup_styles()

    def _setup_styles(self):
        """Setup reusable cell styles"""
        self.header_font = Font(name='Calibri', size=11, bold=True, color=self.colors['header_text'])
        self.header_fill = PatternFill(start_color=self.colors['header_bg'], end_color=self.colors['header_bg'], fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.title_font = Font(name='Calibri', size=16, bold=True)
        self.subtitle_font = Font(name='Calibri', size=12, bold=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def build_report(self, report_content: Dict[str, Any]) -> str:
        """
        Build complete Excel report from content

        Args:
            report_content: Report data from ReportGenerator

        Returns:
            Path to generated Excel file
        """
        logger.info(f"Building Excel report: {report_content['report_id']}")

        # Build worksheets
        self._build_summary_sheet(report_content)
        self._build_campaigns_sheet(report_content)
        self._build_insights_sheet(report_content)
        self._build_recommendations_sheet(report_content)
        self._build_raw_data_sheet(report_content)

        # Save workbook
        self.workbook.save(self.output_path)

        logger.info(f"Excel report generated: {self.output_path}")
        return self.output_path

    def _build_summary_sheet(self, content: Dict[str, Any]):
        """Build executive summary worksheet"""
        ws = self.workbook.create_sheet("Executive Summary", 0)

        row = 1

        # Report title
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"{content['report_type'].replace('_', ' ').title()} Report"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        row += 1

        # Date range
        ws.merge_cells(f'A{row}:F{row}')
        date_range = content.get('date_range', {})
        ws[f'A{row}'].value = f"Period: {date_range.get('start', '')} to {date_range.get('end', '')}"
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        # Company name
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].value = content.get('company_name', 'Your Company')
        ws[f'A{row}'].font = self.subtitle_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Key metrics
        data = content.get('data', {})
        overall = data.get('overall_metrics', {})

        if overall:
            # Metrics header
            ws[f'A{row}'].value = "KEY PERFORMANCE METRICS"
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Create metrics table
            metrics = [
                ('Metric', 'Value'),
                ('Total Spend', f"${overall.get('total_spend', 0):,.2f}"),
                ('Total Revenue', f"${overall.get('total_revenue', 0):,.2f}"),
                ('Return on Ad Spend (ROAS)', f"{overall.get('overall_roas', 0):.2f}x"),
                ('Total Conversions', f"{overall.get('total_conversions', 0):,}"),
                ('Click-Through Rate (CTR)', f"{overall.get('overall_ctr', 0)*100:.2f}%"),
                ('Conversion Rate (CVR)', f"{overall.get('overall_cvr', 0)*100:.2f}%"),
                ('Cost Per Acquisition (CPA)', f"${overall.get('overall_cpa', 0):.2f}"),
                ('Total Impressions', f"{overall.get('total_impressions', 0):,}"),
                ('Total Clicks', f"{overall.get('total_clicks', 0):,}")
            ]

            for i, (label, value) in enumerate(metrics):
                col_offset = 1  # Start from column B
                current_row = row + i

                # Label cell
                label_cell = ws.cell(row=current_row, column=col_offset)
                label_cell.value = label
                label_cell.font = Font(bold=(i == 0))
                label_cell.border = self.border

                # Value cell
                value_cell = ws.cell(row=current_row, column=col_offset + 1)
                value_cell.value = value
                value_cell.font = Font(bold=(i == 0))
                value_cell.border = self.border
                value_cell.alignment = Alignment(horizontal='right')

                # Header row formatting
                if i == 0:
                    label_cell.fill = self.header_fill
                    label_cell.font = self.header_font
                    value_cell.fill = self.header_fill
                    value_cell.font = self.header_font
                elif i % 2 == 0:
                    label_cell.fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid')
                    value_cell.fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid')

            row += len(metrics) + 2

        # Executive summary text
        ws[f'A{row}'].value = "EXECUTIVE SUMMARY"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        summary_text = content.get('summary', 'No summary available.')
        # Remove markdown formatting for Excel
        summary_text = summary_text.replace('**', '').replace('*', '')

        ws.merge_cells(f'A{row}:F{row+5}')
        summary_cell = ws[f'A{row}']
        summary_cell.value = summary_text
        summary_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

    def _build_campaigns_sheet(self, content: Dict[str, Any]):
        """Build detailed campaigns worksheet"""
        ws = self.workbook.create_sheet("Campaign Details")

        data = content.get('data', {})
        campaigns = data.get('campaigns', [])

        if not campaigns:
            ws['A1'].value = "No campaign data available"
            return

        # Headers
        headers = [
            'Campaign ID', 'Campaign Name', 'Status', 'Impressions',
            'Clicks', 'Conversions', 'CTR', 'CVR', 'Spend', 'ROAS', 'CPA',
            'Daily Budget', 'Lifetime Budget'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Data rows
        for row_idx, campaign in enumerate(campaigns, start=2):
            row_data = [
                campaign.get('id', 'N/A'),
                campaign.get('name', 'N/A'),
                campaign.get('status', 'unknown'),
                campaign.get('total_impressions', 0),
                campaign.get('total_clicks', 0),
                campaign.get('total_conversions', 0),
                campaign.get('ctr', 0),
                campaign.get('cvr', 0),
                campaign.get('total_spend', 0),
                campaign.get('avg_roas', 0),
                campaign.get('cpa', 0),
                campaign.get('budget_daily', 0),
                campaign.get('budget_lifetime', 0)
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.border

                # Format numbers
                if col_idx in [4, 5, 6]:  # Impressions, clicks, conversions
                    cell.number_format = '#,##0'
                elif col_idx in [7, 8]:  # CTR, CVR
                    cell.number_format = '0.00%'
                elif col_idx in [9, 11, 12, 13]:  # Spend, CPA, budgets
                    cell.number_format = '$#,##0.00'
                elif col_idx == 10:  # ROAS
                    cell.number_format = '0.00'

                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid')

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Add chart (top campaigns by spend)
        if len(campaigns) >= 2:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Top Campaigns by Spend"
            chart.y_axis.title = 'Spend ($)'
            chart.x_axis.title = 'Campaign'

            # Data range
            data_ref = Reference(ws, min_col=9, min_row=1, max_row=min(11, len(campaigns) + 1))
            cats = Reference(ws, min_col=2, min_row=2, max_row=min(11, len(campaigns) + 1))

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            # Position chart
            ws.add_chart(chart, f"O2")

    def _build_insights_sheet(self, content: Dict[str, Any]):
        """Build insights worksheet"""
        ws = self.workbook.create_sheet("Insights")

        insights = content.get('insights', [])

        row = 1

        # Title
        ws[f'A{row}'].value = "KEY INSIGHTS"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        if not insights:
            ws[f'A{row}'].value = "No insights available"
            return

        # Headers
        headers = ['Type', 'Title', 'Description', 'Metric']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        row += 1

        # Insights
        for insight in insights:
            insight_type = insight.get('type', 'info').upper()
            title = insight.get('title', 'N/A')
            description = insight.get('description', 'N/A')
            metric = insight.get('metric', 'N/A')

            cells = [
                ws.cell(row=row, column=1, value=insight_type),
                ws.cell(row=row, column=2, value=title),
                ws.cell(row=row, column=3, value=description),
                ws.cell(row=row, column=4, value=metric)
            ]

            # Color code by type
            type_colors = {
                'POSITIVE': self.colors['positive'],
                'NEGATIVE': self.colors['negative'],
                'WARNING': self.colors['warning'],
                'INFO': 'FF2196F3'
            }
            type_color = type_colors.get(insight_type, 'FFCCCCCC')

            cells[0].fill = PatternFill(start_color=type_color, end_color=type_color, fill_type='solid')
            cells[0].font = Font(bold=True, color='FFFFFFFF')

            for cell in cells:
                cell.border = self.border
                if row % 2 == 0:
                    if cells.index(cell) > 0:  # Don't override type cell color
                        cell.fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid')

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15

    def _build_recommendations_sheet(self, content: Dict[str, Any]):
        """Build recommendations worksheet"""
        ws = self.workbook.create_sheet("Recommendations")

        recommendations = content.get('recommendations', [])

        row = 1

        # Title
        ws[f'A{row}'].value = "ACTION RECOMMENDATIONS"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 2

        if not recommendations:
            ws[f'A{row}'].value = "No recommendations available"
            return

        # Headers
        headers = ['Priority', 'Title', 'Description', 'Impact', 'Effort']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        row += 1

        # Recommendations
        for rec in recommendations:
            priority = rec.get('priority', 'medium').upper()
            title = rec.get('title', 'N/A')
            description = rec.get('description', 'N/A')
            impact = rec.get('impact', 'N/A')
            effort = rec.get('effort', 'N/A')

            cells = [
                ws.cell(row=row, column=1, value=priority),
                ws.cell(row=row, column=2, value=title),
                ws.cell(row=row, column=3, value=description),
                ws.cell(row=row, column=4, value=impact),
                ws.cell(row=row, column=5, value=effort)
            ]

            # Priority colors
            priority_colors = {
                'HIGH': 'FFD32F2F',
                'MEDIUM': 'FFF57C00',
                'LOW': 'FF388E3C'
            }
            priority_color = priority_colors.get(priority, 'FF757575')

            cells[0].fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type='solid')
            cells[0].font = Font(bold=True, color='FFFFFFFF')
            cells[0].alignment = Alignment(horizontal='center')

            for cell in cells:
                cell.border = self.border
                if row % 2 == 0:
                    if cells.index(cell) > 0:  # Don't override priority cell color
                        cell.fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid')

            # Wrap text for description
            cells[2].alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12

    def _build_raw_data_sheet(self, content: Dict[str, Any]):
        """Build raw data worksheet for custom analysis"""
        ws = self.workbook.create_sheet("Raw Data")

        ws['A1'].value = "This worksheet contains raw data for custom pivot tables and analysis"
        ws['A1'].font = Font(italic=True)
        ws.merge_cells('A1:F1')

        row = 3

        # Export all campaign data in flat format
        data = content.get('data', {})
        campaigns = data.get('campaigns', [])

        if not campaigns:
            ws['A3'].value = "No raw data available"
            return

        # Get all unique keys from campaigns
        all_keys = set()
        for campaign in campaigns:
            all_keys.update(campaign.keys())

        headers = sorted(list(all_keys))

        # Headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        row += 1

        # Data rows
        for campaign in campaigns:
            for col, key in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                value = campaign.get(key)

                # Convert datetime objects to strings
                if isinstance(value, datetime):
                    value = value.isoformat()

                cell.value = value
                cell.border = self.border

                if row % 2 == 0:
                    cell.fill = PatternFill(start_color=self.colors['alt_row'], end_color=self.colors['alt_row'], fill_type='solid')

            row += 1

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15


def generate_excel_report(report_content: Dict[str, Any], output_path: str) -> str:
    """
    Generate Excel report from content

    Args:
        report_content: Report data from ReportGenerator
        output_path: Path to save Excel file

    Returns:
        Path to generated Excel file
    """
    builder = ExcelBuilder(output_path)
    return builder.build_report(report_content)
